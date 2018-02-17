#! /usr/bin/env python


import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


EXCEL_FILE = 'Descr.xlsx'

SEED = 5
M, P, C, D1, D2 = (SEED + i for i in range(5))

# D1:- Description 1
# D2:- Description 2
# M:- Manufacturer
# P:- Product
# C:- Colour

rows = int(input('How many rows? '))

wb = load_workbook(EXCEL_FILE)
ws = wb.active

m = ''
p = ''
c = ''

i = 1
repeated = False
while i < rows:
    i += 1
    print(i)
    
    if i > 2 and ws.cell(row=i, column=P).value == p:
        if repeated:
            continue
        
        else:
            val_D2 = ws.cell(row=i-1, column=D1).value
            val_D1 = ws.cell(row=i-1, column=D2).value
            new_c = ws.cell(row=i, column=C).value
            val_D1 = val_D1.replace(c, new_c)
            val_D2 = val_D2.replace(c, new_c)
            
            cell = ws.cell(row=i, column=D1)
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            cell.font = Font(name='Calibri', size=8)
            cell.value = val_D1
            
            cell = ws.cell(row=i, column=D2)
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            cell.font = Font(name='Calibri', size=8)
            cell.value = val_D2
            
            '''
            for x in (D1, D2):
                cell = ws.cell(row=i, column=x)
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
                cell.font = Font(name='Calibri', size=8)
                cell.value = val_D1 if x == D1 else val_D2
            '''
        
            repeated = True
    
    else:
        repeated = False
        
        if (ws.cell(row=i, column=D1).value == 'SKIP'
            or ws.cell(row=i, column=D2).value == None):
            continue
            
        m = ws.cell(row=i, column=M).value
        p = ws.cell(row=i, column=P).value
        c = ws.cell(row=i, column=C).value
        # m, p, c = (ws.cell(row=i, column=j).value for j in (M, P, C))
        val_D2 = ws.cell(row=i, column=D2).value
        start_ = f"From {m} comes the {p} in {c} colour, "
        end_ = re.search(r'featuring.*$', val_D2).group(0)
        description = start_ + end_
        
        featuring = re.search(r'featuring ([^.]*)[.]', description).group(1)
        sporting = re.search(r'sports? (.*)[.]$', description).group(1)
        description = description.replace(sporting, 'SPORTING')
        description = description.replace(featuring, sporting)
        description = description.replace('SPORTING', featuring)
        
        cell = ws.cell(row=i, column=D1)
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        cell.font = Font(name='Calibri', size=8)
        cell.value = description

wb.save(EXCEL_FILE)
