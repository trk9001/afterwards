"""The new fillsh.py (in development).

This script is intended to eventually replace fillsh.py, since that one hasn't
been updated in ages. Trying to reduce the number of *hacks* present in that
script. Please install `openpyxl` via pip before trying to use this script.
"""

import os.path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class FillSheet:

    DEFAULT_FILE = 'Descr.xlsx'

    def __init__(self, file=DEFAULT_FILE, rows=None):
        """Set up all the required variables.

        Args:
            file -- Name of the Excel file
            rows -- Range of rows to work on

        Raises:
            FileNotFoundError
            TypeError
            ValueError
        """

        self.file = None
        self.rows = None
        self.seed = None

        # Validate and set the file name
        if '.xlsx' not in file:
            raise TypeError('INVALID FILE: '
                            'Must be a valid Excel file ending in .xlsx')
        elif not os.path.exists(file):
            raise FileNotFoundError('FILE DOES NOT EXIST')
        else:
            self.file = file

        # A struct for the `self.rows` data
        class _Rows:
            pass

        # Validate and set the range of rows
        if rows is None:
            self.rows = _Rows()
            self.rows.start = 1
            self.rows.end = self._get_rows_from_file(self.file)
        elif isinstance(rows, str) and re.match(r'^\d*:\d*$', rows):
            self.rows = _Rows()
            row_start, row_end = tuple(rows.split(':'))
            self.rows.start = 1 if row_start == '' else int(row_start)
            self.rows.end = 0 if row_end == '' else int(row_end)
            max_rows = self._get_rows_from_file(self.file)
            if self.rows.end == 0 or self.rows.end > max_rows:
                self.rows.end = max_rows
            if self.rows.start > self.rows.end:
                raise ValueError('INVALID VALUE FOR ROWS: '
                                 'START must be less than END')
        else:
            raise TypeError('INVALID TYPE FOR ROWS: '
                            'Must be empty or of the form START:END')

        # Set the seed column index
        self.seed = self._get_manufacturer_column_index(self.file)
        if self.seed is None:
            raise ValueError('MANUFACTURER COLUMN\'S INDEX NOT FOUND')

    @staticmethod
    def _get_rows_from_file(file):
        """Return the number of rows in a worksheet."""
        wb = load_workbook(file)
        return len(list(wb.active.rows))

    @staticmethod
    def _get_manufacturer_column_index(file):
        """Return the index of the Manufacturer column in a worksheet."""
        wb = load_workbook(file)
        ws = wb.active

        # The Manufacturer column is usually among the first ones,
        # barring the very first column.
        for col in range(2, 10):
            if ws.cell(row=1, column=col).value == 'MANUFACTURER':
                return col
