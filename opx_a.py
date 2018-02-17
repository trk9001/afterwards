#! /usr/bin/env python


from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


EXCEL_FILE = 'Descr.xlsx'

SEED = 5
M, P, C, D1, D2 = (SEED + i for i in range(5))

# D1:- Description 1
# D2:- Description 2
# M:- Manufacturer
# P:- Product
# C:- Colour

rows = int(input('How many rows? '))

wb = load_workbook(EXCEL_FILE)
ws = wb.active

m = ''
p = ''
c = ''

i = 1
while i < rows:
    i += 1
    print(i)
    
    if i > 2 and ws.cell(row=i, column=P).value == p:
        continue
        
    m = ws.cell(row=i, column=M).value
    p = ws.cell(row=i, column=P).value
    c = ws.cell(row=i, column=C).value
    description = f"The {p} from {m} comes in {c} colour, featuring"
    
    cell = ws.cell(row=i, column=D2)
    cell.alignment = Alignment(horizontal='left', wrap_text=True)
    cell.font = Font(name='Calibri', size=8)
    cell.value = description

wb.save(EXCEL_FILE)
