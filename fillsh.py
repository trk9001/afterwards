#! /usr/bin/env python3

"""Script to help fill out product descriptions for a men's fashion store.

When run from the command line, this script takes one required argument to
specify the operation to be performed ('1' for step_1 or '2' for step_2). By
default, it looks for an Excel file named Descr.xlsx in the script's own
directory and processes all non-empty rows. Optionally, it may take a path to an
Excel file (ending with .xlsx), and with the -r option, the number of rows to
process.

Usage examples:
> python -m fillsh a
> ./fillsh.py b /home/user/this.xlsx -r 42

Source & documentation: <https://github.com/trk9001/trkopx>
© 2018 trk9001 <dev.trk.9001@gmail.com>
CC BY-SA 4.0

This project is licensed under a Creative Commons Attribution-ShareAlike 4.0 International License.

"""

import argparse
import os.path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class FillSheet(object):
    """The class that does all the work.

    Attributes:

        DEFAULT_SHEET: Default file to open.
        SHEET: File used by current instance.
        rows: Number of rows in the sheet.
        seed: Starting column index.

    Column labels in the step_1 & step_2 methods:

        D1: Description 1.
        D2: Description 2.
        M: Manufacturer.
        P: Product.
        C: Colour.

    """

    DEFAULT_SHEET = 'Descr.xlsx'

    def __init__(self, sheet=None, rows=None):
            """Constructor.

            Raises:

                TypeError: Invalid file, must have extension .xlsx.
                TypeError: Invalid type for rows in constructor.
                TypeError: Not seeded from file.
                FileNotFoundError

            """

            if sheet:
                if '.xlsx' in sheet:
                    self.SHEET = sheet
                else:
                    raise TypeError('Invalid file, must have extension .xlsx.')

            else:
                self.SHEET = self.DEFAULT_SHEET

            if not os.path.exists(self.SHEET):
                raise FileNotFoundError

            if rows is None:
                self.rows = self.get_rows()

            elif isinstance(rows, int):
                self.rows = rows

            else:
                raise TypeError('Invalid type for rows in constructor.')

            seed = self.get_mc()

            if seed:
                self.seed = seed

            else:
                raise TypeError('Not seeded from file.')

    def get_rows(self):
            """Return the number of rows in the sheet."""

            wb = load_workbook(self.SHEET)
            return len(list(wb.active.rows))

    def get_mc(self):
            """Return the index of the Manufacturer column."""

            wb = load_workbook(self.SHEET)
            ws = wb.active

            for col in range(2, 10):
                if ws.cell(row=1, column=col).value == 'MANUFACTURER':
                    return col

    def step_1(self, rows=None):
            """Initial operation.

            Args:
                rows: Optional argument for the number of rows to process.

            Raises:
                TypeError: Invalid type for rows in step_1.

            """

            if rows is None:
                rows = self.rows

            elif not isinstance(rows, int):
                raise TypeError('Invalid type for rows in step_1.')

            # Set column numbers sequentially from seed value
            M, P, C, D1, D2 = (self.seed + i for i in range(5))

            wb = load_workbook(self.SHEET)
            ws = wb.active

            p = ''

            i = 1
            while i < rows:
                i += 1

                # Handler for cases where there is some text in the second
                # description column.
                if ws.cell(row=i, column=D2).value:
                    tmp = ws.cell(row=i, column=D1).value
                    if tmp is None:
                        tmp = ''

                    tmp = tmp + '; ' + ws.cell(row=i, column=D2).value
                    ws.cell(row=i, column=D1).value = tmp
                    ws.cell(row=i, column=D2).value = None

                if i > 2 and ws.cell(row=i, column=P).value == p:
                    continue

                m = ws.cell(row=i, column=M).value
                p = ws.cell(row=i, column=P).value
                c = ws.cell(row=i, column=C).value

                descr = f'The {p} from {m} comes in {c} colour, featuring'

                self.format_cell(ws.cell(row=i, column=D2))
                ws.cell(row=i, column=D2).value = descr

            wb.save(self.SHEET)

    def step_2(self, rows=None):
            """Final operation.

            Args:
                rows: Optional argument for the number of rows to process.

            Raises:
                TypeError: Invalid type for rows in step_2.

            """

            if rows is None:
                rows = self.rows

            elif not isinstance(rows, int):
                raise TypeError('Invalid type for rows in step_2.')

            # Set column numbers sequentially from seed value
            M, P, C, D1, D2 = (self.seed + i for i in range(5))

            wb = load_workbook(self.SHEET)
            ws = wb.active

            m = ''
            p = ''
            c = ''

            i = 1
            times_repeated = 0
            while i < rows:
                i += 1

                # Manual skip condition (either cell works)
                if (ws.cell(row=i, column=D1).value == 'SKIP'
                        or ws.cell(row=i, column=D2).value == 'SKIP'):
                    continue

                if ws.cell(row=i, column=P).value == p:  # Last p (not updated yet)
                    if times_repeated == 0 or times_repeated == 2:
                        new_c = ws.cell(row=i, column=C).value

                        new_D1 = ws.cell(row=i-1, column=D2).value
                        new_D1 = new_D1.replace(c, new_c)

                        new_D2 = ws.cell(row=i-1, column=D1).value
                        new_D2 = new_D2.replace(c, new_c)

                    elif times_repeated == 1:
                        new_c = ws.cell(row=i, column=C).value

                        elder = f'From {m} comes'
                        youngster = f'{m} offers'
                        new_D1 = ws.cell(row=i-2, column=D1).value
                        new_D1 = new_D1.replace(elder, youngster).replace(c, new_c)

                        elder = f'The {p} from {m}'
                        youngster = f'Offered by {m}, the {p}'
                        new_D2 = ws.cell(row=i-2, column=D2).value
                        new_D2 = new_D2.replace(elder, youngster).replace(c, new_c)

                    else:
                        continue

                    self.format_cell(ws.cell(row=i, column=D1))
                    ws.cell(row=i, column=D1).value = new_D1

                    self.format_cell(ws.cell(row=i, column=D2))
                    ws.cell(row=i, column=D2).value = new_D2

                    times_repeated += 1

                else:
                    times_repeated = 0
                    full_descr = True

                    if ws.cell(row=i, column=D1).value == 'EZPZ':
                        full_descr = False

                    m = ws.cell(row=i, column=M).value
                    p = ws.cell(row=i, column=P).value
                    c = ws.cell(row=i, column=C).value

                    D2_val = ws.cell(row=i, column=D2).value
                    start_ = f'From {m} comes the {p} in {c} colour, '
                    end_ = ''
                    try:
                        end_ = re.search(r'featuring.*$', D2_val).group(0)
                    except Exception as e:
                        print(i, type(e), e)

                    # TODO(trk9001) Fix possible pre-assignment references
                    descr = start_ + end_
                    if full_descr:
                        try:
                            ft = re.search(r'featuring ([^.]*)[.]', descr).group(1)
                            sport = re.search(r'sports? (.*)[.]$', descr).group(1)
                        except Exception as e:
                            print(i, type(e), e)

                        descr = descr.replace(sport, '$')
                        descr = descr.replace(ft, sport)
                        descr = descr.replace('$', ft)

                    self.format_cell(ws.cell(row=i, column=D1))
                    ws.cell(row=i, column=D1).value = descr

            wb.save(self.SHEET)

    @staticmethod
    def format_cell(cell):

            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            cell.font = Font(name='Calibri', size=8)

# End of FillSheet


def main():
    """Execute the CLI argument parser."""

    parser = argparse.ArgumentParser(
            usage='%(prog)s [-h] <OPERATION> [FILE] [-r R]',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            description=('Script to help fill out product descriptions\n'
                         "for a men's fashion store."),
            epilog=('Source: https://github.com/trk9001/trkopx\n'
                    '© 2018 trk9001 <dev.trk.9001@gmail.com>\n'
                    'CC BY-SA 4.0')
    )

    parser.add_argument(
            'OPERATION',
            choices=['1', '2'],
            help='choose one of the two OPERATIONS (step 1 or 2)'
    )

    parser.add_argument(
            'FILE',
            nargs='?',
            help='path to the Excel file (Descr.xlsx)'
    )

    parser.add_argument(
            '-r', '--rows',
            dest='R',
            type=int,
            help='process R rows (defaults to all)'
    )

    args = parser.parse_args()

    fs = FillSheet(args.FILE, args.R)
    getattr(fs, 'step_' + args.OPERATION)()


if __name__ == '__main__':
    main()
