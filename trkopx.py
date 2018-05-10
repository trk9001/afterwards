"""Script to help fill out product descriptions for Stand-Out.net.

Under development. Since all the necessary methods are encapsulated in a class,
it may be imported (ie., `from trkopx import FillSheet`) for use interactively.

Prerequisites (subject to changes):
> Have file Descr.xlsx in the same directory as this script.

Usage:
> Instantiate FillSheet with or without the number of rows as a parameter.
> Invoke alpha() without or with the number of rows as a parameter.
> Invoke beta() without or with the number of rows as a parameter.
> Pay attention to and handle any raised exceptions.

```
from trkopx import FillSheet
o = FillSheet(69)
o.alpha()
o.beta()
```
"""

import argparse
import os.path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class FillSheet:

    SHEET = 'Descr.xlsx'

    def __init__(self, rows=None):
        if not os.path.exists(self.SHEET):
            raise FileNotFoundError

        if rows is None or isinstance(rows, int):
            self.rows = rows
        else:
            raise TypeError('Invalid type for rows in constructor')

        self.seed = self.get_mc()
        if self.seed is None:
            raise TypeError('Not seeded from file')

        self.a = self.alpha
        self.b = self.beta

    @classmethod
    def get_mc(cls):
        """Return the number of the Manufacturer column"""

        wb = load_workbook(cls.SHEET)
        ws = wb.active

        for col in range(2, 10):
            if ws.cell(row=1, column=col).value == 'MANUFACTURER':
                return col

    @staticmethod
    def format_cell(cell):
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
        cell.font = Font(name='Calibri', size=8)

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # Column labels in the alpha & beta methods:
    #
    # D1: Description 1
    # D2: Description 2
    # M: Manufacturer
    # P: Product
    # C: Colour
    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    def alpha(self, rows=None):
        if rows is not None:
            if isinstance(rows, int):
                self.rows = rows
            else:
                raise TypeError('Invalid type for rows in alpha')

        if self.rows is None:
            raise TypeError('Invalid type for self.rows in alpha')

        # Set column numbers sequentially from seed value
        M, P, C, D1, D2 = (self.seed + i for i in range(5))

        wb = load_workbook(self.SHEET)
        ws = wb.active

        p = ''

        i = 1
        while i < self.rows:
            i += 1

            if i > 2 and ws.cell(row=i, column=P).value == p:
                continue

            m = ws.cell(row=i, column=M).value
            p = ws.cell(row=i, column=P).value
            c = ws.cell(row=i, column=C).value

            descr = f'The {p} from {m} comes in {c} colour, featuring'
            # TODO: Refine to accommodate plural products

            self.format_cell(ws.cell(row=i, column=D2))
            ws.cell(row=i, column=D2).value = descr

        wb.save(self.SHEET)

    def beta(self, rows=None):
        if rows is not None:
            if isinstance(rows, int):
                self.rows = rows
            else:
                raise TypeError('Invalid type for rows in beta')

        if self.rows is None:
            raise TypeError('Invalid type for self.rows in beta')

        # Set column numbers sequentially from seed value
        M, P, C, D1, D2 = (self.seed + i for i in range(5))

        wb = load_workbook(self.SHEET)
        ws = wb.active

        m = ''
        p = ''
        c = ''

        i = 1
        times_repeated = 0
        while i < self.rows:
            i += 1

            # Manual skip condition (either cell works)
            if (ws.cell(row=i, column=D1).value == 'SKIP'
                    or ws.cell(row=i, column=D2).value == 'SKIP'):
                continue

            if ws.cell(row=i, column=P).value == p:  # Last p (not updated yet)
                if times_repeated == 0 or times_repeated == 2:
                    new_c = ws.cell(row=i, column=C).value

                    new_D1 = ws.cell(row=i-1, column=D2).value
                    new_D1 = new_D1.replace(c, new_c)

                    new_D2 = ws.cell(row=i-1, column=D1).value
                    new_D2 = new_D2.replace(c, new_c)

                elif times_repeated == 1:
                    new_c = ws.cell(row=i, column=C).value

                    elder = f'From {m} comes'
                    youngster = f'{m} offers'
                    new_D1 = ws.cell(row=i-2, column=D1).value
                    new_D1 = new_D1.replace(elder, youngster).replace(c, new_c)

                    elder = f'The {p} from {m}'
                    youngster = f'Offered by {m}, the {p}'
                    new_D2 = ws.cell(row=i-2, column=D2).value
                    new_D2 = new_D2.replace(elder, youngster).replace(c, new_c)

                else:
                    continue

                self.format_cell(ws.cell(row=i, column=D1))
                ws.cell(row=i, column=D1).value = new_D1

                self.format_cell(ws.cell(row=i, column=D2))
                ws.cell(row=i, column=D2).value = new_D2

                times_repeated += 1

            else:
                times_repeated = 0
                full_descr = True

                if ws.cell(row=i, column=D1).value == 'EZPZ':
                    full_descr = False

                m = ws.cell(row=i, column=M).value
                p = ws.cell(row=i, column=P).value
                c = ws.cell(row=i, column=C).value

                D2_val = ws.cell(row=i, column=D2).value
                start_ = f'From {m} comes the {p} in {c} colour, '
                end_ = ''
                try:
                    end_ = re.search(r'featuring.*$', D2_val).group(0)
                except Exception as e:
                    print(i, type(e), e)
                # TODO: Refine to accommodate plural products

                descr = start_ + end_
                if full_descr:
                    try:
                        ft = re.search(r'featuring ([^.]*)[.]', descr).group(1)
                        sport = re.search(r'sports? (.*)[.]$', descr).group(1)
                    except Exception as e:
                        print(i, type(e), e)

                    descr = descr.replace(sport, '#')
                    descr = descr.replace(ft, sport)
                    descr = descr.replace('#', ft)

                self.format_cell(ws.cell(row=i, column=D1))
                ws.cell(row=i, column=D1).value = descr

        wb.save(self.SHEET)

# End of FillSheet


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
